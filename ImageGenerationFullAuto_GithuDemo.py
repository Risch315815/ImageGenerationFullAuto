# Install these package with pip:  [requests, json, time, hashlib, openpyxl, shutil] and [win11toast]
# 先用 pip 安裝這些套件： [requests, json, time, hashlib, openpyxl, shutil] and [win11toast]
# For operative system before Windows 11, please find and delete every line that contains "toast"
# 作業系統版本比 Windows 11 舊的使用者，可以刪除所有包含 "toast" 的行 (ctrl+f 尋找)

# Get an account on TensorArt (https://tensor.art/login), then get your API keys and App_ID at (https://tams.tensor.art/app)
# 先到 (https://tensor.art/login) 創帳號，然後在 (https://tams.tensor.art/app) 取得 API 金鑰和 App_ID

#--------------------------------------------------------------
# Personal information #通關密碼
API_Key = str(Your API key) # Type is String #資料型態是：字串
App_ID = str(Your App ID) # Type is String #資料型態是：字串

# If you want to generate image from image, set to  "True"; if set to "False", it would generate only from prompt
# 如果你想要從圖片生成圖片，請設為 "True"；如果設為 "False"，則會從提示詞生成圖片
generate_from_image = True
# Img source directory #圖片來源路徑(請從 C槽/D槽/E槽 開始)
img_source_path = {Path to storage-directory 資料夾路徑} #for example, C:\Image_Source
# Please put the source image in the directory with the same name as the row number (row-number-directory),
# then put the row-number-directory under the storage-directory 
# 請將放置圖片的資料夾命名為 Excel檔案 中存放提示詞那列的編號，然後將該資料夾放在 image_source_path這個資料夾裡面
# (for example, C:\Image_Source\4)

# About directory where you put your image: for each image in the directory: 
# this program will make a sub-directory with the same name as the image, and put the image into the sub-directory, 
# then put generated image into that sub-directory
# 關於來源圖片應該怎麼放：storage-directory內的每張圖片都會建立一個子資料夾，子資料夾名稱同圖片名稱，
# 然後將圖片會被複製到子資料夾裡面，然後生成的圖片會存在子資料夾裡面
# (for example, C:\Image_Source\4\image_001.jpg => C:\Image_Source\4\image_001\image_001.png [生成圖片會存在這個資料夾裡面])


# Please put your prompt in an Excel file, and then put the Excel file's path here
# 請將提示詞放在 Excel 檔案裡面，然後將 Excel 檔案的路徑放在這裡
promt_source = {Path to Excel file Excel檔案路徑}
this_book_tab = {Name of tab in Excel file Excel檔案內書籤名稱}
this_row_list = [1, 3, 4...] # Put as many row numbers as you want to generate #這是 list，所以想要文生圖/圖生圖的Excel行編號都可以加進來

how_many_image = 60 # How many pictures do you want to generate? #每組提示詞想生成幾張圖
img_width = 1024 # Maximum is 1024 #圖片寬度最大為 1024
img_height = 1024 # Maximum is 1024 #圖片高度最大為 1024
img_gen_steps = 30 # Maximum is 60, suggest 30 #圖片生成步數最大為 60，建議 30
cgf_scale = 8 # How much image generation accord with prompt; maximum is 14, suggest 7-10 #提示詞的圖片生成比例，最大為 14，建議 7-10

# Model #模型
model_01 = {
    "Name of model":"serial number of model",
    "模型名稱":"模型序號"
    }

#--------------------------------------------------------------

class TensorAPI():
    
    def __init__(self, prompt_dict, excell, storage):
        self.prompt_dict = prompt_dict
        self.excell = excell
        self.storage = storage
        # url pack
        self.url_pre = "https://ap-east-1.tensorart.cloud"
        self.app_id = App_ID
        self.url_local = "http://localhost:8080"
        self.url_workflow = '/v1/workflows'
        self.url_job = "/v1/jobs"
        self.url_resource = "/v1/resource"
        #parameter pack
        self.def_furry = self.prompt_dict['def_furry']
        self.species = self.prompt_dict['species']
        self.body_type = self.prompt_dict['body_type']
        self.participant = self.prompt_dict['participant']
        self.activity = self.prompt_dict['activity']
        self.extremities = self.prompt_dict['extremities']
        self.wearing = self.prompt_dict['wearing']
        self.facial_expression = self.prompt_dict['facial_expression']
        self.environment = self.prompt_dict['environment']
        self.shot_angle = self.prompt_dict['shot_angle']
        self.lighting = self.prompt_dict['lighting']
        self.art_style = self.prompt_dict['art_style']
        self.quality = self.prompt_dict['quality']
        self.misc = self.prompt_dict['misc']

#--------------------------------------------------------------

    def pending_img_full_auto(self, directory, num_of_images, SD_model):
        import os
        import shutil
        import time
        from win11toast import toast

        file_list = next(os.walk(directory))[2] # [2]: file
        print(file_list)
        for file in file_list:
            img_format_list = ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp'] 
            # You can add more image format here #你可以在這裡加入更多種圖片格式
            for img_format in img_format_list:
                if img_format in file:
                    dir_name = file.replace(img_format, '')
                    temp_dir_path = directory + '\\' + dir_name
                    if not os.path.exists(temp_dir_path):
                        os.mkdir(temp_dir_path)
                    else:
                        for i in range(1, 100):
                            temp_dir_path = directory + '\\' + dir_name+ f'_{i}'
                            if os.path.exists(temp_dir_path):
                                continue
                            else:
                                os.mkdir(temp_dir_path)
                                break


                    src = directory + f'\\{file}'
                    dst = temp_dir_path + f'\\{dir_name}.png'
                    shutil.copyfile(src, dst)
                    print(f'{dir_name} copied')
                else:
                    continue

        desig_img = []
        img_dir_list = next(os.walk(directory))[1] # [1]: directory
        for img_dir in img_dir_list:
            inside_dir = directory + f'\\{img_dir}'
            if os.path.isdir(inside_dir): # check if it is a directory, for 'Shortcut.lnk'
                if not next(os.walk(inside_dir))[1]: # check if the list is without directory
                    if next(os.walk(inside_dir))[2]: # check if the list is without file
                        desig_img.append(img_dir)
                else:
                    pass
        print(desig_img)

        for img_dir in desig_img:
            for i in range(num_of_images):
                pending_img_dir = directory + f'\\{img_dir}'
                pending_img_path = directory + f'\\{img_dir}\\{next(os.walk(directory + f'\\{img_dir}'))[2][0]}'
                
                start_time = time.time()
                self.img_gen(pending_img_path, pending_img_dir, SD_model)
                end_time = time.time()
                execution_time = end_time - start_time
                print(f"Execution time: {execution_time} seconds")
                time.sleep(2)
            toast(f'Image {img_dir} from source {directory[-2:]} is done')

#--------------------------------------------------------------

    # IMG2IMG: IMG source ID
    def upload_img(self, img_path):
        import json
        import requests

        print(img_path)
        data = {"expireSec": 3600}

        response = requests.post(f"{self.url_pre}{self.url_resource}/image", json=data, headers={
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': f'Bearer {API_Key}'
        })
        print(response.text)
        response_data = json.loads(response.text)
        resource_id = response_data['resourceId']
        put_url = response_data['putUrl']
        headers = response_data['headers']
        with open(img_path, 'rb') as f:
            res = f.read()
            response = requests.put(put_url, data=res, headers=headers)
            print(response.text)
        return resource_id

#--------------------------------------------------------------

    def excel_to_dicts(excell, which_row, book_tab):
        import openpyxl
        from openpyxl import load_workbook

        wb = openpyxl.load_workbook(excell)
        sheet = wb[book_tab]
        headers = [cell.value for cell in sheet[1]]
        dictionaries = {}
        for row in sheet.iter_rows(min_row=which_row, max_row=which_row, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            dictionaries.update(row_dict) 
        return dictionaries

#--------------------------------------------------------------

    def img_gen(self, img_path, storage, SD_model):
        import hashlib
        import time

        image_source = {}
        if generate_from_image == True:
            image_source = {"image_resource_id": f"{self.upload_img(img_path)}"}
        else:
            image_source = {"seed": -1}

        img2img_data = {
        "request_id": hashlib.md5(str(int(time.time())).encode()).hexdigest(),
        "stages": [
            {
                "type": "INPUT_INITIALIZE",
                "inputInitialize": image_source.update({"count": 1})
                },
            {
                "type": "DIFFUSION",
                "diffusion": {
                    "width": img_width,
                    "height": img_height,
                    "prompts": [
                        {
                            "text": f"""
                            {self.def_furry}, {self.species}, 
                            {self.body_type},   
                            {self.participant}, {self.activity},  
                            {self.extremities},
                            {self.wearing}, 
                            {self.facial_expression}, 
                            {self.environment}, {self.shot_angle}, {self.lighting}, 
                            {self.art_style}, 
                            {self.quality},
                            {self.misc}
                            """
                        }
                    ],
                    
                    "negativePrompts": [
                {
                    "text": """
                    (deformityv6, bad-image-v2, bad-hands-5, EasynegativeV2, verybadimagenegative_v1.3), 
                    (out of frame), (blurry, blurry eyes, missing pupils), (bulge), (body fusion), 
                    (missing leg: 1.4), (extra leg:1.4), (extra head: 1.4), (extra arm: 1.4), (minor:1.4)
                    """
                }
            ],

            "sampler": "DPM++ 2M Karras",
            "sdVae": "pastel-waifu-diffusion.vae.pt",
            "steps": img_gen_steps,
            "sd_model": SD_model,
            "clip_skip": 1,
            "cfg_scale": cgf_scale
        
                }
            }
        ]
    }

        response_data = self.create_job(img2img_data)
        if 'job' in response_data:
            job_dict = response_data['job']
            job_id = job_dict.get('id')
            job_status = job_dict.get('status')
            print(job_id, job_status)
            temp_job = self.get_job_result(job_id)
            print('temp_job = ',temp_job)
            self.store_img(temp_job, job_id, storage)
            #storage should be the local directory where the image is stored, string!!!


    def create_job(self, data):
        import requests
        import json

        response = requests.post(f"{self.url_pre}{self.url_job}", json=data, headers={
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'Authorization': f'Bearer {API_Key}'
        })
        print(response.text)
        return json.loads(response.text)


    def get_job_result(self, job_id):
        import time
        import requests
        import json

        nnn = 0
        while True:
            nnn += 1
            time.sleep(1)
            response = requests.get(f"{self.url_pre}{self.url_job}/{job_id}", headers={
                'Content-Type': 'application/json',
                'Accept': 'application/json',
                'Authorization': f'Bearer {API_Key}'
            })
            get_job_response_data = json.loads(response.text)
            if 'job' in get_job_response_data:
                job_dict = get_job_response_data['job']
                job_status = job_dict.get('status')
                if job_status == 'SUCCESS':
                    print(job_dict)
                    break
                elif job_status == 'FAILED':
                    print(job_dict)
                    break
                else:
                    print(job_dict)
        return job_dict


    def store_img(self, job_dict, job_id, local_storage):
        import time
        import os
        import requests

        local_time = time.gmtime()[:11]
        img_url = job_dict['successInfo']['images'][0]['url']
        img_response = requests.get(img_url)
        local_time = time.strftime("%Y_%m_%d", time.localtime())
        print(local_time)
        img_dir = local_storage+f"\\{local_time}"
        if not os.path.exists(img_dir):
            os.makedirs(img_dir)
        with open(img_dir+f'\\{job_id}.png', 'wb') as f:
            f.write(img_response.content)

#--------------------------------------------------------------

# Execution
if __name__ == '__main__':

    import time

    for this_row in this_row_list:
        for model in model_01:
            img_source_dir = f'{img_source_path}\\{str(this_row)}'
            param = TensorAPI.excel_to_dicts(promt_source, this_row, this_book_tab)
            print(param)
            api = TensorAPI(param, promt_source, img_source_dir)

            total_start_time = time.time()
            api.pending_img_full_auto(img_source_dir, how_many_image, model)
            total_end_time = time.time()
            total_execution_time = total_end_time - total_start_time
            print(f"Total Execution Time: {total_execution_time} seconds")
